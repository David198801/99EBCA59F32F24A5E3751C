# coding:utf-8
import openpyxl
import xlwt
import os

#函数，按日分割xlsx，写入xls
def clipXlsx(file, readPath, savePath):
    book = openpyxl.load_workbook(readPath + file)#读xlsx
    for month in range(1, 13):#遍历1到12月
        sheet = book[file[-9:-5] + str(month).zfill(2)]#读取当月sheet

        # 创建工作簿及sheet
        wbook = xlwt.Workbook()
        wsheet = wbook.add_sheet("sheet")

        # 函数，写入第0行
        def writeFirst():
            wsheet.write(0, 0, 'latitude')
            wsheet.write(0, 1, 'longitude')
            wsheet.write(0, 2, 'elevation')
            wsheet.write(0, 3, 'temperature')

        writeFirst()

        rows = sheet.max_row #获取最大行数
        dayCheck = 1 #用于检查日期，从1日开始
        row_xls = 0 #用于记录写入行数，由于写入前行数递增，从0开始
        # 保存文件后需重置行数及写入，为避免重复写入第二行，行数递增在写入之前

        for row_num in range(1, rows + 1): #遍历当前sheet每一行
            #当前行各项值
            latitude = sheet.cell(row=row_num, column=2).value
            longitude = sheet.cell(row=row_num, column=3).value
            elevation = sheet.cell(row=row_num, column=4).value
            dayNum = sheet.cell(row=row_num, column=5).value
            temperature = sheet.cell(row=row_num, column=6).value

            #函数，写入各项值
            def writeData():
                wsheet.write(row_xls, 0, latitude)
                wsheet.write(row_xls, 1, longitude)
                wsheet.write(row_xls, 2, elevation)
                wsheet.write(row_xls, 3, temperature)

            #函数，构造文件名，保存文件，输出文件名
            def saveData():
                #文件名=地名（读取目录切片）+日期（文件名切片）+月份 +日期+ 后缀名
                saveName = readPath[7:-1] + file[-9:-5] + str(month).zfill(2) + str(dayCheck).zfill(2) + ".xls"
                wbook.save(savePath + saveName)
                print saveName

            #整个写入过程
            if dayNum is None: #跳过空行
                pass
            elif row_num == rows: #最后一行处理，由于排序，最后一行可能为数据或者表头
                # 若日期列值为数字，则改行为数据行，写入行数+1，写入行，保存文件
                if isinstance(dayNum, (int, long)):
                    row_xls += 1
                    writeData()
                    saveData()
                else: #否（日期列为汉字）则直接保存文件
                    saveData()
            elif dayCheck == dayNum: #检查日期，若相同则写入行数+1，写入行
                row_xls += 1
                writeData()
            else: #否（日期不一致，且非空行且非最后一行）则
                saveData() #保存文件
                dayCheck = dayNum #更新日期

                #新建工作簿及sheet，写入表头行
                wbook = xlwt.Workbook()
                wsheet = wbook.add_sheet("sheet")
                writeFirst()
                #重置写入行数，从第二行开始，写入第二行
                row_xls = 1
                writeData()


readPath = "D:/air/HuBei/"
savePath = "D:/air/day_xls/"

fileList = os.listdir(readPath) #列出目录下所有文件
#遍历文件，判断xlsx，执行函数
for file in fileList:
    if file[-5:] == ".xlsx":
        clipXlsx(file, readPath, savePath)
