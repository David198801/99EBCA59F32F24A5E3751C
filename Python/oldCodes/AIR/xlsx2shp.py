# coding:utf-8
import os
import openpyxl
import shapefile

path = 'C:/Users/admin/Desktop/aaa/HuBei'

def shpTans(xlsFile):
    wb = openpyxl.load_workbook(xlsFile)
    sheet = wb[xlsFile[-9:-5]]

    shp = shapefile.Writer(shapeType=1)
    month = [u'一月', u'二月', u'三月', u'四月', u'五月', u'六月', u'七月', u'八月', u'九月', u'十月', u'十一月', u'十二月']
    for m in month:
        shp.field(m.encode("utf8"), 'C', '10')

    rows = sheet.max_row
    cols = sheet.max_column

