import requests
from bs4 import BeautifulSoup
import re
import openpyxl

pageNumAll = 2
#name = "柔音馆"
name = "溢音馆"
pageUrlPart = "https://rainkmc.com/forum.php?mod=forumdisplay&fid=42&orderby=dateline&filter=author&orderby=dateline&page="


headers = {"Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3",
"Accept-Encoding":"gzip, deflate, br",
"Accept-Language":"zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6,ja;q=0.5",
"Cache-Control":"max-age=0",
"Connection":"keep-alive",
"Cookie":"ihFD_2132_smile=9D1; ihFD_2132_nofavfid=1; ihFD_2132_saltkey=Tbx6By22; ihFD_2132_lastvisit=1589379673; ihFD_2132_auth=25a6l%2FZZaevEHMmg6zBAnTY2n9S4yt4IwIuCcs78zqPkcr5LmqupNyEDKXRyEb8Z1bKO4vHLz3d3cqmPVOBcm%2Bpn; ihFD_2132_lastcheckfeed=5094%7C1589383298; ihFD_2132_myrepeat_rr=R0; ihFD_2132_sid=EVSSuz; ihFD_2132_lip=140.240.24.125%2C1589950062; ihFD_2132_punchindex=0; ihFD_2132_punchtime=1589904000; ihFD_2132_ulastactivity=a178IoiDhUhRFtPWJuFTi6pVDy3Zj67oQqlbLN%2Fgp0BdyudzyFcA; ihFD_2132_st_t=5094%7C1589950444%7C7d05cb7f18f28ce861ab9417cb62c417; ihFD_2132_forum_lastvisit=D_38_1583893443D_43_1584753936D_37_1585466540D_58_1585618768D_57_1587699013D_56_1588212575D_46_1588481488D_2_1589105717D_42_1589802058D_40_1589902670D_41_1589950444; ihFD_2132_sendmail=1; ihFD_2132_lastact=1589950446%09misc.php%09patch",
"Host":"rainkmc.com",
"Referer":"https//rainkmc.com/",
"Sec-Fetch-Mode":"navigate",
"Sec-Fetch-Site":"same-origin",
"Sec-Fetch-User":"?1",
"Upgrade-Insecure-Requests":"1",
"User-Agent":"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.108 Safari/537.36"}


wb = openpyxl.Workbook()
ws = wb.active
ws.title = name
ws.cell(row=1, column=1).value = "RJ号"
ws.cell(row=1, column=2).value = "标题"
ws.cell(row=1, column=3).value = "链接"
rowNum = 1

for pageNum in range(1,pageNumAll+1):
    pageUrl = pageUrlPart + str(pageNum)
    print(pageNum,"/",pageNumAll)
    print(pageUrl)
    while True:
        try:
            r = requests.get(pageUrl,headers=headers,timeout=10)
            break
        except Exception as e:
            print(e)
            continue
    bs = BeautifulSoup(r.text,features="lxml")
    postTable = bs.find_all("table",id="threadlisttableid")[0]
    postList = postTable.find_all("tbody",id=re.compile(r"normalthread_\d+"))
    for post in postList:
        postALabel = post.tr.th.find_all("a",class_="s xst")[0]
        postTitle = postALabel.string
        postUrl = "https://rainkmc.com/" + postALabel["href"]
        rjCodeList = re.findall(r"\w{2}\d{6}",postTitle,re.I)

        print(rjCodeList,postTitle)

        rowNum += 1
        if not rjCodeList:
            pass
        elif len(rjCodeList)>1:
            rjCode = ""
            for i in rjCodeList:
                rjCode += i + " "
            rjCode=rjCode[:-1]
            ws.cell(row=rowNum, column=1).value = rjCode
        else:
            ws.cell(row=rowNum, column=1).value = rjCodeList[0]
        ws.cell(row=rowNum, column=2).value = postTitle
        ws.cell(row=rowNum, column=3).value = postUrl
        #ws.cell(row=rowNum, column=3).value = '=HYPERLINK("{}", "{}")'.format(postUrl, "点击打开")



wb.save("D:\\linshi\\"+name+".xlsx")